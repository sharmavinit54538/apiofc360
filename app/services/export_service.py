"""Export service layer for handling data exports in Excel, CSV, and PDF formats.

Ensures proper multi-tenant isolation, database-driven filtering, and audit logging.
"""

from __future__ import annotations

import csv
import io
import json
import logging
import uuid
from datetime import date, datetime, timezone
from decimal import Decimal
from typing import Any, Sequence

from fastapi import status
from sqlalchemy import select, and_, or_, func
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm import selectinload

from app.core.exceptions import AppException, BadRequestException, DatabaseException
from app.attendance.models.attendance import Attendance
from app.models.employee import Employee
from app.models.department import Department
from app.models.manager import Manager
from app.models.payroll import PayrollAttendanceInput, Payslip, PayrollRun
from app.models.employee_leave_policy import EmployeeLeavePolicy
from app.models.performance import PerformanceReview, PerformanceReviewCycle
from app.models.audit_log import AuditLog
from app.models.company import Company
from app.models.user import User

try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill
    from openpyxl.utils import get_column_letter
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

try:
    from reportlab.lib.pagesizes import letter, landscape
    from reportlab.lib import colors
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.pdfgen import canvas
    HAS_REPORTLAB = True
except ImportError:
    HAS_REPORTLAB = False

logger = logging.getLogger(__name__)


if HAS_REPORTLAB:
    # Custom canvas class to compute total page numbers dynamically
    class NumberedCanvas(canvas.Canvas):
        def __init__(self, *args, **kwargs):
            super().__init__(*args, **kwargs)
            self._saved_page_states = []

        def showPage(self):
            self._saved_page_states.append(dict(self.__dict__))
            self._startPage()

        def save(self):
            num_pages = len(self._saved_page_states)
            for state in self._saved_page_states:
                self.__dict__.update(state)
                self.draw_page_number(num_pages)
                super().showPage()
            super().save()

        def draw_page_number(self, page_count):
            self.saveState()
            self.setFont("Helvetica", 9)
            self.setFillColor(colors.HexColor("#64748B"))
            # Page dimensions: letter is 8.5 x 11 inches, landscape is 11 x 8.5 inches
            page_width, page_height = self._pagesize
            
            # Draw header separator
            self.setStrokeColor(colors.HexColor("#E2E8F0"))
            self.setLineWidth(0.5)
            self.line(36, page_height - 40, page_width - 36, page_height - 40)
            
            # Draw header title
            self.drawString(36, page_height - 32, "Ecochange HRMS Report")
            self.drawRightString(page_width - 36, page_height - 32, datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC"))
            
            # Draw footer line
            self.line(36, 45, page_width - 36, 45)
            
            # Draw footer page count
            page_text = f"Page {self._pageNumber} of {page_count}"
            self.drawRightString(page_width - 36, 30, page_text)
            self.drawString(36, 30, "Confidential - For Internal Use Only")
            self.restoreState()
else:
    class NumberedCanvas:
        pass


class ExportService:
    def __init__(self, session: AsyncSession) -> None:
        self.session = session

    # ------------------------------------------------------------------
    # Data Fetchers
    # ------------------------------------------------------------------

    async def _fetch_employees(self, company_id: uuid.UUID, filters: dict[str, Any]) -> list[dict[str, Any]]:
        stmt = select(Employee).where(and_(Employee.company_id == company_id, Employee.is_deleted == False))
        
        if filters.get("department"):
            stmt = stmt.where(Employee.department == filters["department"])
        if filters.get("status"):
            stmt = stmt.where(Employee.status == filters["status"].upper())
        if filters.get("employee_type"):
            stmt = stmt.where(Employee.employment_type == filters["employee_type"].upper())
        if filters.get("designation"):
            stmt = stmt.where(Employee.designation == filters["designation"])
        if filters.get("search"):
            pattern = f"%{filters['search']}%"
            stmt = stmt.where(or_(
                Employee.first_name.ilike(pattern),
                Employee.last_name.ilike(pattern),
                Employee.employee_id.ilike(pattern),
                Employee.personal_email.ilike(pattern)
            ))
            
        stmt = stmt.order_by(Employee.created_at.desc())
        res = await self.session.execute(stmt)
        rows = res.scalars().all()
        
        return [
            {
                "ID": r.employee_id,
                "First Name": r.first_name,
                "Last Name": r.last_name,
                "Email": r.personal_email,
                "Phone": r.phone,
                "Department": r.department,
                "Designation": r.designation,
                "Type": r.employment_type,
                "Capacity": getattr(r, "employee_capacity", 100) or 100,
                "Cost Center": getattr(r, "cost_center_id", "") or "",
                "Status": r.status,
                "Joining Date": r.joining_date.isoformat() if r.joining_date else "",
                "Shift": r.shift or "General"
            }
            for r in rows
        ]

    async def _fetch_departments(self, company_id: uuid.UUID, filters: dict[str, Any]) -> list[dict[str, Any]]:
        stmt = select(Department).where(Department.company_id == company_id)
        
        if filters.get("status"):
            stmt = stmt.where(Department.status == filters["status"].upper())
        if filters.get("search"):
            pattern = f"%{filters['search']}%"
            stmt = stmt.where(or_(
                Department.department_name.ilike(pattern),
                Department.department_code.ilike(pattern)
            ))
            
        stmt = stmt.order_by(Department.created_at.desc())
        res = await self.session.execute(stmt)
        rows = res.scalars().all()
        
        return [
            {
                "Code": r.department_code,
                "Name": r.department_name,
                "Description": r.description or "",
                "Location": r.location or "Headquarters",
                "Capacity": getattr(r, "employee_capacity", 100) or 100,
                "Status": r.status,
                "Created At": r.created_at.date().isoformat() if r.created_at else ""
            }
            for r in rows
        ]

    async def _fetch_managers(self, company_id: uuid.UUID, filters: dict[str, Any]) -> list[dict[str, Any]]:
        stmt = select(Manager).where(and_(Manager.company_id == company_id, Manager.is_deleted == False))
        
        if filters.get("department"):
            stmt = stmt.where(Manager.department == filters["department"])
        if filters.get("status"):
            stmt = stmt.where(Manager.status == filters["status"].upper())
        if filters.get("search"):
            pattern = f"%{filters['search']}%"
            stmt = stmt.where(or_(
                Manager.first_name.ilike(pattern),
                Manager.last_name.ilike(pattern),
                Manager.manager_id.ilike(pattern),
                Manager.personal_email.ilike(pattern)
            ))
            
        stmt = stmt.order_by(Manager.joining_date.desc())
        res = await self.session.execute(stmt)
        rows = res.scalars().all()
        
        return [
            {
                "ID": r.manager_id,
                "First Name": r.first_name,
                "Last Name": r.last_name,
                "Email": r.personal_email,
                "Phone": r.phone,
                "Department": r.department,
                "Designation": r.designation,
                "Status": r.status,
                "Joining Date": r.joining_date.isoformat() if r.joining_date else ""
            }
            for r in rows
        ]

    async def _fetch_attendance(self, company_id: uuid.UUID, filters: dict[str, Any]) -> list[dict[str, Any]]:
        stmt = select(PayrollAttendanceInput, Employee).join(Employee, PayrollAttendanceInput.employee_id == Employee.id).where(PayrollAttendanceInput.company_id == company_id)
        
        if filters.get("search"):
            pattern = f"%{filters['search']}%"
            stmt = stmt.where(or_(
                Employee.first_name.ilike(pattern),
                Employee.last_name.ilike(pattern),
                Employee.employee_id.ilike(pattern)
            ))
        if filters.get("date_from"):
            try:
                dt = datetime.fromisoformat(filters["date_from"])
                stmt = stmt.where(PayrollAttendanceInput.period_year >= dt.year)
            except ValueError:
                pass
                
        stmt = stmt.order_by(PayrollAttendanceInput.period_year.desc(), PayrollAttendanceInput.period_month.desc())
        res = await self.session.execute(stmt)
        rows = res.all()
        
        return [
            {
                "Employee ID": emp.employee_id,
                "Employee Name": f"{emp.first_name} {emp.last_name}",
                "Period": f"{att.period_year}-{att.period_month:02d}",
                "Paid Days": float(att.paid_days),
                "LOP Days": float(att.lop_days),
                "Arrears": float(att.arrears),
                "One-Time Bonus": float(att.one_time_bonus),
                "Remarks": att.remarks or ""
            }
            for att, emp in rows
        ]

    async def _fetch_face_attendance(self, company_id: uuid.UUID, filters: dict[str, Any]) -> list[dict[str, Any]]:
        stmt = select(Attendance).join(Employee, Attendance.employee_id == Employee.id).where(Attendance.company_id == company_id)
        
        if filters.get("search"):
            pattern = f"%{filters['search']}%"
            stmt = stmt.where(or_(
                Employee.first_name.ilike(pattern),
                Employee.last_name.ilike(pattern),
                Employee.employee_id.ilike(pattern)
            ))
        if filters.get("branch") and filters["branch"].lower() not in {"", "all"}:
            stmt = stmt.where(Employee.branch == filters["branch"])
        if filters.get("department") and filters["department"].lower() not in {"", "all"}:
            stmt = stmt.where(Employee.department == filters["department"])
        if filters.get("date_from"):
            try:
                dt = datetime.fromisoformat(filters["date_from"])
                stmt = stmt.where(Attendance.date >= dt.date())
            except ValueError:
                pass
                
        stmt = stmt.options(selectinload(Attendance.employee)).order_by(Attendance.date.desc(), Attendance.check_in_time.desc())
        res = await self.session.execute(stmt)
        rows = res.scalars().all()
        
        return [
            {
                "Employee ID": r.employee.employee_id if r.employee else "",
                "Employee Name": f"{r.employee.first_name} {r.employee.last_name}" if r.employee else "Unknown",
                "Date": r.date.isoformat(),
                "Check-In Time": r.check_in_time.isoformat() if r.check_in_time else "",
                "Check-Out Time": r.check_out_time.isoformat() if r.check_out_time else "",
                "Working Hours": float(r.working_hours) if r.working_hours is not None else 0.0,
                "IP Address": r.ip_address or "",
                "Device Info": r.device_info or "",
                "GPS Coordinates": f"{r.latitude},{r.longitude}" if r.latitude is not None else "",
                "Check-In Photo": r.face_image_url or "",
                "Check-Out Photo": r.checkout_image_url or "",
            }
            for r in rows
        ]

    async def _fetch_leaves(self, company_id: uuid.UUID, filters: dict[str, Any]) -> list[dict[str, Any]]:
        stmt = select(EmployeeLeavePolicy).join(Employee).where(Employee.company_id == company_id)
        
        if filters.get("search"):
            pattern = f"%{filters['search']}%"
            stmt = stmt.where(or_(
                Employee.first_name.ilike(pattern),
                Employee.last_name.ilike(pattern),
                Employee.employee_id.ilike(pattern)
            ))
        if filters.get("status"):  # leave type
            stmt = stmt.where(EmployeeLeavePolicy.leave_type == filters["status"])
            
        stmt = stmt.order_by(EmployeeLeavePolicy.leave_type.asc())
        res = await self.session.execute(stmt.options(selectinload(EmployeeLeavePolicy.employee)))
        rows = res.scalars().all()
        
        return [
            {
                "Employee ID": r.employee.employee_id,
                "Employee Name": f"{r.employee.first_name} {r.employee.last_name}",
                "Leave Type": r.leave_type,
                "Total Allocated Days": float(r.total_days),
                "Used Days": float(r.used_days),
                "Carry Forward": "Yes" if r.carry_forward else "No",
                "Effective From": r.effective_from.isoformat() if r.effective_from else "",
                "Effective To": r.effective_to.isoformat() if r.effective_to else ""
            }
            for r in rows
        ]

    async def _fetch_payroll(self, company_id: uuid.UUID, filters: dict[str, Any]) -> list[dict[str, Any]]:
        stmt = select(Payslip).join(Employee).where(Payslip.company_id == company_id)
        
        if filters.get("status"):
            stmt = stmt.where(Payslip.payment_status == filters["status"].upper())
        if filters.get("search"):
            pattern = f"%{filters['search']}%"
            stmt = stmt.where(or_(
                Employee.first_name.ilike(pattern),
                Employee.last_name.ilike(pattern),
                Employee.employee_id.ilike(pattern),
                Payslip.payslip_number.ilike(pattern)
            ))
            
        stmt = stmt.order_by(Payslip.period_year.desc(), Payslip.period_month.desc())
        res = await self.session.execute(stmt.options(selectinload(Payslip.employee)))
        rows = res.scalars().all()
        
        return [
            {
                "Payslip No": r.payslip_number,
                "Employee ID": r.employee.employee_id,
                "Employee Name": f"{r.employee.first_name} {r.employee.last_name}",
                "Period": f"{r.period_year}-{r.period_month:02d}",
                "Basic": float(r.basic),
                "HRA": float(r.hra),
                "Gross Earnings": float(r.gross_earnings),
                "Total Deductions": float(r.total_deductions),
                "Net Pay": float(r.net_pay),
                "Status": r.payment_status,
                "Payment Date": r.payment_date.isoformat() if r.payment_date else ""
            }
            for r in rows
        ]

    async def _fetch_performance(self, company_id: uuid.UUID, filters: dict[str, Any]) -> list[dict[str, Any]]:
        stmt = select(PerformanceReview).join(Employee).where(Employee.company_id == company_id)
        
        if filters.get("status"):
            stmt = stmt.where(PerformanceReview.status == filters["status"].upper())
        if filters.get("search"):
            pattern = f"%{filters['search']}%"
            stmt = stmt.where(or_(
                Employee.first_name.ilike(pattern),
                Employee.last_name.ilike(pattern),
                Employee.employee_id.ilike(pattern)
            ))
            
        stmt = stmt.order_by(PerformanceReview.created_at.desc())
        res = await self.session.execute(stmt.options(
            selectinload(PerformanceReview.employee),
            selectinload(PerformanceReview.cycle)
        ))
        rows = res.scalars().all()
        
        return [
            {
                "Cycle": r.cycle.name,
                "Employee ID": r.employee.employee_id,
                "Employee Name": f"{r.employee.first_name} {r.employee.last_name}",
                "Self Rating": float(r.self_rating) if r.self_rating is not None else "",
                "Reviewer Rating": float(r.reviewer_rating) if r.reviewer_rating is not None else "",
                "AI Score": float(r.ai_overall_score) if r.ai_overall_score is not None else "",
                "AI Recommendation": "Promotion Recommended" if r.promotion_recommendation else "No Change",
                "Increment %": float(r.salary_increment_percentage) if r.salary_increment_percentage is not None else 0.0,
                "Status": r.status
            }
            for r in rows
        ]

    # ------------------------------------------------------------------
    # Audit Log Writer
    # ------------------------------------------------------------------

    async def _write_audit_log(
        self,
        user_id: uuid.UUID,
        company_id: uuid.UUID,
        module: str,
        filters: dict[str, Any],
        fmt: str,
        ip_address: str | None,
        user_agent: str | None,
    ) -> None:
        try:
            user = await self.session.get(User, user_id)
            email = user.email if user else None
            
            details = {
                "company_id": str(company_id),
                "export_module": module,
                "filters_used": {k: str(v) for k, v in filters.items() if v is not None},
                "export_format": fmt,
            }
            
            log = AuditLog(
                id=uuid.uuid4(),
                user_id=user_id,
                action="EXPORT",
                email=email,
                ip_address=ip_address,
                user_agent=user_agent or "HRMS Portal",
                details=json.dumps(details),
                created_at=datetime.now(timezone.utc),
            )
            self.session.add(log)
            await self.session.flush()
        except Exception as e:
            logger.error("Failed to write export audit log: %s", str(e), exc_info=True)

    # ------------------------------------------------------------------
    # Renderers
    # ------------------------------------------------------------------

    def _render_csv(self, data: list[dict[str, Any]], headers: list[str]) -> bytes:
        output = io.StringIO()
        writer = csv.DictWriter(output, fieldnames=headers)
        writer.writeheader()
        for row in data:
            writer.writerow(row)
        return output.getvalue().encode("utf-8")

    def _render_xlsx(self, data: list[dict[str, Any]], headers: list[str], title: str) -> bytes:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = title[:30]  # Excel tab name max 31 chars
        
        # Styling definitions
        font_header = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
        fill_header = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid") # Dark Slate Gray
        font_cell = Font(name="Segoe UI", size=10)
        align_left = Alignment(horizontal="left", vertical="center")
        align_right = Alignment(horizontal="right", vertical="center")
        
        # Write headers
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = font_header
            cell.fill = fill_header
            cell.alignment = align_left
            
        # Write data rows
        for row_idx, row_data in enumerate(data, 2):
            for col_idx, header in enumerate(headers, 1):
                val = row_data.get(header)
                cell = ws.cell(row=row_idx, column=col_idx, value=val)
                cell.font = font_cell
                
                # Check data type for alignment & formats
                if isinstance(val, (int, float)):
                    cell.alignment = align_right
                    if isinstance(val, float):
                        cell.number_format = "#,##0.00"
                else:
                    cell.alignment = align_left
                    
        # Auto-fit column widths with padding
        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                val_str = str(cell.value or "")
                if len(val_str) > max_len:
                    max_len = len(val_str)
            ws.column_dimensions[col_letter].width = max(max_len + 4, 12)
            
        buffer = io.BytesIO()
        wb.save(buffer)
        return buffer.getvalue()

    def _render_pdf(
        self,
        data: list[dict[str, Any]],
        headers: list[str],
        title: str,
        company_name: str,
        generated_by: str,
    ) -> bytes:
        # Create bytes buffer
        buffer = io.BytesIO()
        
        # Landscape document setup
        doc = SimpleDocTemplate(
            buffer,
            pagesize=landscape(letter),
            rightMargin=36,
            leftMargin=36,
            topMargin=54,
            bottomMargin=54,
        )
        
        styles = getSampleStyleSheet()
        
        # Custom stylesheet definitions
        style_title = ParagraphStyle(
            name="ReportTitle",
            parent=styles["Heading1"],
            fontName="Helvetica-Bold",
            fontSize=20,
            leading=24,
            textColor=colors.HexColor("#0F172A"),
            spaceAfter=6,
        )
        
        style_meta = ParagraphStyle(
            name="ReportMeta",
            parent=styles["Normal"],
            fontName="Helvetica",
            fontSize=9,
            leading=13,
            textColor=colors.HexColor("#475569"),
            spaceAfter=15,
        )
        
        style_table_header = ParagraphStyle(
            name="TableHeader",
            parent=styles["Normal"],
            fontName="Helvetica-Bold",
            fontSize=8,
            leading=10,
            textColor=colors.white,
        )
        
        style_table_cell = ParagraphStyle(
            name="TableCell",
            parent=styles["Normal"],
            fontName="Helvetica",
            fontSize=8,
            leading=10,
            textColor=colors.HexColor("#334155"),
        )

        elements = []
        
        # 1. Report Header info
        elements.append(Paragraph(f"{company_name} — {title}", style_title))
        meta_text = (
            f"<b>Generated By:</b> {generated_by} | "
            f"<b>Generated Date:</b> {datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M UTC')} | "
            f"<b>Total Records:</b> {len(data)}"
        )
        elements.append(Paragraph(meta_text, style_meta))
        
        # 2. Build table grid
        table_data = []
        # Header row
        table_data.append([Paragraph(h, style_table_header) for h in headers])
        # Data rows
        for row in data:
            row_cells = []
            for h in headers:
                val = row.get(h)
                val_str = str(val) if val is not None else ""
                row_cells.append(Paragraph(val_str, style_table_cell))
            table_data.append(row_cells)
            
        # Standard landscape width is 792 points. Margin left+right = 72, usable = 720.
        col_width = 720.0 / len(headers)
        
        t = Table(table_data, colWidths=[col_width] * len(headers))
        
        # Alternating styling grid
        t_style = [
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1E293B")),
            ("ALIGN", (0, 0), (-1, -1), "LEFT"),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
            ("TOPPADDING", (0, 0), (-1, -1), 4),
            ("LEFTPADDING", (0, 0), (-1, -1), 4),
            ("RIGHTPADDING", (0, 0), (-1, -1), 4),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#E2E8F0")),
        ]
        
        # Alternating row colors
        for idx in range(1, len(data) + 1):
            if idx % 2 == 0:
                t_style.append(("BACKGROUND", (0, idx), (-1, idx), colors.HexColor("#F8FAFC")))
                
        t.setStyle(TableStyle(t_style))
        elements.append(t)
        
        # Build PDF
        doc.build(elements, canvasmaker=NumberedCanvas)
        return buffer.getvalue()

    # ------------------------------------------------------------------
    # Export Engine
    # ------------------------------------------------------------------

    async def export_module(
        self,
        *,
        user_id: uuid.UUID,
        company_id: uuid.UUID,
        module: str,
        filters: dict[str, Any],
        fmt: str = "xlsx",
        ip_address: str | None = None,
        user_agent: str | None = None,
    ) -> tuple[bytes, str, str]:
        """Main export gateway logic. Fetches, formats, logs, and returns file content."""
        
        # Validate format
        fmt = fmt.lower()
        if fmt not in {"xlsx", "csv", "pdf"}:
            raise BadRequestException("Invalid export format. Supported formats are: xlsx, csv, pdf.")

        # Load company profile for headers/meta
        company = await self.session.get(Company, company_id)
        company_name = company.name if company else "Ecochange"

        user = await self.session.get(User, user_id)
        generated_by = user.name if user else "System Admin"

        # Module router to fetch data
        if module == "employees":
            data = await self._fetch_employees(company_id, filters)
            headers = [
                "ID", "First Name", "Last Name", "Email", "Phone",
                "Department", "Designation", "Type", "Capacity", "Cost Center", "Status", "Joining Date", "Shift"
            ]
            title = "Employees Report"
            filename_prefix = "employees"
        elif module == "departments":
            data = await self._fetch_departments(company_id, filters)
            headers = ["Code", "Name", "Description", "Location", "Capacity", "Status", "Created At"]
            title = "Departments Report"
            filename_prefix = "departments"
        elif module == "managers":
            data = await self._fetch_managers(company_id, filters)
            headers = ["ID", "First Name", "Last Name", "Email", "Phone", "Department", "Designation", "Status", "Joining Date"]
            title = "Managers Report"
            filename_prefix = "managers"
        elif module == "attendance":
            data = await self._fetch_attendance(company_id, filters)
            headers = ["Employee ID", "Employee Name", "Period", "Paid Days", "LOP Days", "Arrears", "One-Time Bonus", "Remarks"]
            title = "Attendance Input Report"
            filename_prefix = "attendance"
        elif module == "face_attendance":
            data = await self._fetch_face_attendance(company_id, filters)
            headers = [
                "Employee ID", "Employee Name", "Date", "Check-In Time", "Check-Out Time",
                "Working Hours", "IP Address", "Device Info", "GPS Coordinates", "Check-In Photo", "Check-Out Photo"
            ]
            title = "Face Attendance Report"
            filename_prefix = "face_attendance"
        elif module == "leaves":
            data = await self._fetch_leaves(company_id, filters)
            headers = ["Employee ID", "Employee Name", "Leave Type", "Total Allocated Days", "Used Days", "Carry Forward", "Effective From", "Effective To"]
            title = "Leaves Allocation Report"
            filename_prefix = "leaves"
        elif module == "payroll":
            data = await self._fetch_payroll(company_id, filters)
            headers = ["Payslip No", "Employee ID", "Employee Name", "Period", "Basic", "HRA", "Gross Earnings", "Total Deductions", "Net Pay", "Status", "Payment Date"]
            title = "Payroll Payslip Report"
            filename_prefix = "payroll"
        elif module == "performance":
            data = await self._fetch_performance(company_id, filters)
            headers = ["Cycle", "Employee ID", "Employee Name", "Self Rating", "Reviewer Rating", "AI Score", "AI Recommendation", "Increment %", "Status"]
            title = "Performance Reviews Report"
            filename_prefix = "performance"
        else:
            raise AppException(message=f"Export module '{module}' not found.", status_code=status.HTTP_404_NOT_FOUND)

        # Log action to audit logs
        await self._write_audit_log(
            user_id=user_id,
            company_id=company_id,
            module=module,
            filters=filters,
            fmt=fmt,
            ip_address=ip_address,
            user_agent=user_agent
        )
        await self.session.commit()

        # Render file bytes
        today = date.today().isoformat()
        if fmt == "csv":
            content = self._render_csv(data, headers)
            filename = f"{filename_prefix}_{today}.csv"
            media_type = "text/csv"
        elif fmt == "pdf":
            content = self._render_pdf(data, headers, title, company_name, generated_by)
            filename = f"{filename_prefix}_{today}.pdf"
            media_type = "application/pdf"
        else:  # default: xlsx
            content = self._render_xlsx(data, headers, title)
            filename = f"{filename_prefix}_{today}.xlsx"
            media_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

        return content, filename, media_type
